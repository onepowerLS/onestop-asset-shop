#!/usr/bin/env python3
"""
Extract RET Material Request Log from docs/ Excel -> JSON for Firestore import.

Reads:  docs/RET Material Items Database and Requests Log (1).xlsx
        Sheet: "RET items request log"
Writes: migration/output/am_core_archived_requests.json

Usage:  python3 migration/etl_archived_ret_requests.py
"""

import json
import re
from datetime import datetime, timezone, date
from pathlib import Path

import openpyxl


PROJECT = Path(__file__).resolve().parent.parent
SRC = PROJECT / "docs" / "RET Material Items Database and Requests Log (1).xlsx"
OUT = PROJECT / "migration" / "output" / "am_core_archived_requests.json"
SHEET = "RET items request log"

# All sites in this sheet are Lesotho concessions
SITE_COUNTRY = {
    "MAT": "LSO", "TLH": "LSO", "MAS": "LSO", "SHG": "LSO",
    "KET": "LSO", "SEH": "LSO", "TOS": "LSO", "SEB": "LSO",
    "MAK": "LSO", "HQ": "LSO",
}


def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("none", "null", "n/a", "na", "-", "to verify", "unknown", ""):
        return ""
    return s


def clean_date_str(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = clean(val)
    if not s:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def parse_items(raw):
    """Split free-text items string into a list of cleaned tokens."""
    if not raw:
        return []
    raw = str(raw).strip()
    parts = re.split(r"[\n\r]+", raw)
    out = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if "," in part:
            for sub in part.split(","):
                sub = sub.strip()
                if sub:
                    out.append(sub)
        else:
            out.append(part)
    return out


def extract():
    if not SRC.exists():
        print(f"ERROR: Source file not found: {SRC}")
        return

    wb = openpyxl.load_workbook(SRC, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[SHEET]
    records = []
    seq_by_year = {}

    for r in range(2, ws.max_row + 1):
        ts = ws.cell(r, 1).value
        if ts is None or not isinstance(ts, datetime):
            continue

        requester_email = clean(ws.cell(r, 2).value)
        requester_name = clean(ws.cell(r, 3).value)
        items_raw = clean(ws.cell(r, 4).value)
        site = clean(ws.cell(r, 5).value).upper()
        dispatch_date = ws.cell(r, 6).value
        receiver_name = clean(ws.cell(r, 7).value)
        receiver_email = clean(ws.cell(r, 8).value)
        notes = clean(ws.cell(r, 9).value)
        db_updated = clean(ws.cell(r, 10).value).lower()

        year = ts.year
        seq_by_year.setdefault(year, 0)
        seq_by_year[year] += 1
        seq = seq_by_year[year]

        records.append({
            "archived_request_number": f"RET-{year}-{seq:03d}",
            "timestamp": ts.replace(tzinfo=None).isoformat() + "+02:00",
            "requester_email": requester_email,
            "requester_name": requester_name,
            "items_requested_raw": items_raw,
            "items_requested_list": parse_items(items_raw),
            "site_code": site,
            "site_country_code": SITE_COUNTRY.get(site, "LSO"),
            "estimated_dispatch_date": clean_date_str(dispatch_date),
            "receiver_name": receiver_name,
            "receiver_email": receiver_email,
            "notes": notes,
            "database_updated": db_updated in ("yes", "true", "1"),
            "source_type": "GoogleForm_RET_RequestLog",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(records)} records to {OUT}")

    # Summaries
    sites = {}
    for rec in records:
        s = rec["site_code"]
        sites[s] = sites.get(s, 0) + 1
    print("\nPer site:")
    for s, n in sorted(sites.items()):
        print(f"  {s}: {n}")

    years = {}
    for rec in records:
        y = rec["archived_request_number"][4:8]
        years[y] = years.get(y, 0) + 1
    print("\nPer year:")
    for y, n in sorted(years.items()):
        print(f"  {y}: {n}")


if __name__ == "__main__":
    extract()
