#!/usr/bin/env python3
"""
Import SIM CARD VERIFICATION FINAL.xlsx → JSON for am_core_sim_cards.

Does not call Firestore by default; writes migration/output/am_core_sim_cards.import.json
for review or use with import_to_firestore.py (extend COLLECTIONS if you batch-upload).

Usage:
  python3 migration/import_sim_workbook.py [path/to/SIM\ CARD\ VERIFICATION\ FINAL.xlsx]

Requires: openpyxl (pip install openpyxl)
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
SKIP_SHEETS = {"deactivated sims"}  # import separately if needed


def normalize_msisdn(raw: str) -> str:
    if raw is None:
        return ""
    s = re.sub(r"\D+", "", str(raw).strip())
    return s


def row_to_record(sheet_name: str, row: tuple, header: list[str | None]) -> dict | None:
    def col(name: str) -> str | float | None:
        try:
            i = header.index(name)
        except ValueError:
            return None
        if i >= len(row):
            return None
        v = row[i]
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        return v

    num = col("NUMBER")
    if num is None:
        return None
    norm = normalize_msisdn(str(num))
    if not norm or len(norm) < 6:
        return None

    status_raw = col("STATUS") or col("status") or "Unknown"
    if isinstance(status_raw, str):
        st = status_raw.strip().lower()
        if st in ("active", "activer"):
            status = "Active"
        elif "deactiv" in st:
            status = "Deactivated"
        else:
            status = "Unknown"
    else:
        status = "Active"

    loc = col("SIM LOCATION")
    person = col("PERSON ASSIGNED TO")
    contact = col("CONTACT VALUE")
    locate = col("Locate/Could not locate")
    comments = col("comments")

    notes_parts = []
    if comments:
        notes_parts.append(str(comments))
    aq = col("Actions / Questions")
    if aq:
        notes_parts.append("Actions: " + str(aq))

    now = datetime.now(timezone.utc).isoformat()
    return {
        "msisdn_normalized": norm,
        "msisdn_display": str(num).strip(),
        "contact_value": str(contact) if contact is not None else "",
        "pool": sheet_name.strip(),
        "sim_location": str(loc).strip() if loc is not None else "",
        "person_assigned": str(person).strip() if person is not None else "",
        "status": status,
        "locate_status": str(locate).strip() if locate is not None else "",
        "notes": "\n\n".join(notes_parts),
        "legacy_import": {"workbook_sheet": sheet_name},
        "created_at": now,
        "updated_at": now,
        "created_by": "import_sim_workbook.py",
        "updated_by": "import_sim_workbook.py",
    }


def find_header_row(ws) -> tuple[int, list[str | None]]:
    for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True)):
        cells = [str(c).strip() if c is not None else None for c in row]
        if any(c == "NUMBER" for c in cells if c):
            return i, cells
    return -1, []


def main() -> int:
    default_xlsx = (
        Path(__file__).resolve().parents[1]
        / "SIM CARD VERIFICATION FINAL.xlsx"
    )
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_xlsx
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        hdr_row, header = find_header_row(ws)
        if hdr_row < 0:
            continue
        for row in ws.iter_rows(min_row=hdr_row + 2, values_only=True):
            rec = row_to_record(sheet_name, tuple(row), header)
            if not rec:
                continue
            key = rec["msisdn_normalized"]
            if key in seen:
                continue
            seen.add(key)
            records.append(rec)
        ws.reset_dimensions()  # read_only optimization

    wb.close()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_file = OUTPUT_DIR / "am_core_sim_cards.import.json"
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2)

    print(f"Wrote {len(records)} SIM records to {out_file}")
    print("Review the file, then upload via admin tooling or extend migration/import_to_firestore.py.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
