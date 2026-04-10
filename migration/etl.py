#!/usr/bin/env python3
"""
ETL: Extract → Transform → Load-ready JSON for Firestore migration.

Reads all Dropbox Excel sources, classifies items per the 4-tier model
(FixedAsset, Material, Consumable, Inventory), deduplicates, and outputs
one JSON file per target Firestore collection under migration/output/.

Usage:
    python3 migration/etl.py
"""

import json
import os
import re
import hashlib
from datetime import datetime, date
from pathlib import Path

import openpyxl

DROPBOX = "/Users/mattmso/Dropbox/1PWR"
OUT_DIR = Path(__file__).resolve().parent / "output"
OUT_DIR.mkdir(parents=True, exist_ok=True)

NOW = datetime.utcnow().isoformat() + "Z"

# ── Normalisation helpers ────────────────────────────────────────────

def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("none", "null", "n/a", "na", "-", "to verify", "unknown", ""):
        return ""
    return s

def clean_int(val):
    s = clean(val)
    if not s:
        return 0
    try:
        return max(0, int(float(s)))
    except (ValueError, TypeError):
        return 0

def clean_float(val):
    s = clean(val)
    if not s:
        return 0.0
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return 0.0

def clean_date(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = clean(val)
    if not s:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""

def dedup_key(name, serial, tag):
    raw = f"{name.lower()}|{serial.lower()}|{tag.lower()}"
    return hashlib.md5(raw.encode()).hexdigest()


# ── Category → item_class mapping ───────────────────────────────────

ASSET_DB_CATEGORY_MAP = {
    "IT Equipment":       "FixedAsset",
    "EE Equipment":       "FixedAsset",
    "Coms Equipment":     "FixedAsset",
    "Test Equipment":     "FixedAsset",
    "Plant and Vehicles": "FixedAsset",
    "Machine Tools":      "FixedAsset",
    "Hand Tools":         "FixedAsset",
    "Furniture":          "FixedAsset",
    "Funiture":           "FixedAsset",
    "furniture":          "FixedAsset",
    "Office Equipment":   "FixedAsset",
    "Housing":            "FixedAsset",
    "Storage":            "FixedAsset",
    "Safety":             "Consumable",
    "Accessory":          "Consumable",
    "accessory":          "Consumable",
}

DEPT_MAP = {
    "ADMIN DEPT":        "General",
    "CONSTRUCTION DEPT": "RET",
    "EHS":               "O&M",
    "ENG DEPT":          "O&M",
    "Fleet":             "General",
    "IT":                "General",
    "IT DEPT":           "General",
    "OPS":               "O&M",
    "OPS DEPT":          "O&M",
    "PRODUCTION DEPT":   "RET",
}

CONDITION_MAP = {
    "(2) good": "Good",
    "(4) bad":  "Poor",
}

PPE_CATEGORY_MAP = {
    "215": ("FixedAsset", "FA-BLD", "Buildings"),
    "216": ("FixedAsset", "FA-FUR", "Furniture & Fixtures"),
    "217": ("FixedAsset", "FA-MCH", "Machinery & Equipment"),
    "218": ("FixedAsset", "FA-VEH", "Vehicles"),
    "219": ("FixedAsset", "FA-ELE", "Electronics & Tools"),
}

CONCESSION_COUNTRY = {
    "MAT": "LSO", "TLH": "LSO", "MAS": "LSO", "SHG": "LSO",
    "LEB": "LSO", "RIB": "LSO", "KET": "LSO", "SEH": "LSO",
    "LSB": "LSO", "MAK": "LSO", "TOS": "LSO", "SEB": "LSO",
    "HQ":  "LSO",
    "BENIN": "BEN", "BEN": "BEN",
}


# ── Source 1: Asset Spreadsheet Database ─────────────────────────────

def extract_asset_db():
    path = f"{DROPBOX}/1PWR Asset Management/1_Assets Management/Assets Database/Asset spreadsheet database.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["All Assets_To Access DB"]

    items = []
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 5).value)
        if not name:
            continue

        cat_raw = clean(ws.cell(r, 7).value)
        item_class = ASSET_DB_CATEGORY_MAP.get(cat_raw, "FixedAsset")

        dept_raw = clean(ws.cell(r, 8).value)
        dept = DEPT_MAP.get(dept_raw, "General")

        cond_raw = clean(ws.cell(r, 9).value)
        condition = CONDITION_MAP.get(cond_raw, "Good")

        serial = clean(ws.cell(r, 4).value)
        tag_old = clean(ws.cell(r, 2).value)
        tag_new = clean(ws.cell(r, 3).value)

        price = clean_float(ws.cell(r, 16).value)
        if item_class == "FixedAsset" and price > 0 and price < 200:
            item_class = "Consumable"

        items.append({
            "name": name,
            "description": clean(ws.cell(r, 6).value),
            "item_class": item_class,
            "category_name": cat_raw,
            "department_scope": dept,
            "serial_number": serial,
            "manufacturer": clean(ws.cell(r, 12).value),
            "model": clean(ws.cell(r, 13).value),
            "purchase_date": clean_date(ws.cell(r, 20).value),
            "purchase_price": price,
            "current_value": clean_float(ws.cell(r, 17).value),
            "condition_status": condition,
            "status": "Retired" if clean(ws.cell(r, 15).value) else "Available",
            "quantity": clean_int(ws.cell(r, 14).value) or 1,
            "unit_of_measure": "EA",
            "location_name": clean(ws.cell(r, 11).value),
            "assigned_to": clean(ws.cell(r, 21).value),
            "country_code": "LSO",
            "notes": clean(ws.cell(r, 18).value),
            "legacy_tag": tag_old or tag_new,
            "source": "AssetSpreadsheetDB",
        })
    return items


# ── Source 2: PPE Financial Register ─────────────────────────────────

def extract_ppe_register():
    path = f"{DROPBOX}/OnePower/financial/14 - annual financial statements/March 2025/parts/FY2025 1PWR  PPE ASSET REGISTER _REV.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Calcs"]

    items = []
    current_cat = None
    for r in range(3, ws.max_row + 1):
        cat_code = clean(ws.cell(r, 1).value)
        if cat_code in PPE_CATEGORY_MAP:
            current_cat = cat_code

        raw_name = clean(ws.cell(r, 2).value)
        cost = clean_float(ws.cell(r, 3).value)
        if not raw_name or cost == 0:
            continue

        name = re.sub(r"\s*-?\s*Tag\s*#?\s*.*$", "", raw_name, flags=re.IGNORECASE).strip()
        name = re.sub(r"^\d{5}\s+", "", name).strip()

        tags = re.findall(r"(1PWR\d{5}|[A-Z]{2,10}\d{3,6})", raw_name)

        cat_info = PPE_CATEGORY_MAP.get(current_cat or "217", ("FixedAsset", "FA-MCH", "Machinery"))

        items.append({
            "name": name,
            "description": "",
            "item_class": cat_info[0],
            "category_code": cat_info[1],
            "category_name": cat_info[2],
            "department_scope": "General",
            "purchase_price": cost,
            "current_value": clean_float(ws.cell(r, 14).value),
            "salvage_value": round(cost * 0.05, 2),
            "depreciation_this_year": abs(clean_float(ws.cell(r, 11).value)),
            "condition_status": "Good",
            "status": "Available",
            "quantity": 1,
            "unit_of_measure": "EA",
            "country_code": "LSO",
            "legacy_tags": tags,
            "source": "FY2025_PPE_Register",
        })
    return items


# ── Source 3: Material databases (RET, FAC, O&M, General) ───────────

def _extract_material_sheet(path, sheet, name_col, desc_col, unit_col, qty_proc_col, qty_bal_col, dept, skip_header=1):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    items = []
    for r in range(skip_header + 1, ws.max_row + 1):
        name = clean(ws.cell(r, name_col).value)
        if not name:
            continue
        qty = clean_int(ws.cell(r, qty_bal_col).value) if qty_bal_col else clean_int(ws.cell(r, qty_proc_col).value)
        items.append({
            "name": name,
            "description": clean(ws.cell(r, desc_col).value) if desc_col else "",
            "item_class": "Material",
            "department_scope": dept,
            "quantity": max(qty, 0),
            "unit_of_measure": clean(ws.cell(r, unit_col).value).upper() or "EA",
            "condition_status": "Good",
            "status": "Available" if qty > 0 else "Consumed",
            "country_code": "LSO",
            "source": f"DeptSheet_{dept}",
        })
    return items

def extract_materials():
    base = f"{DROPBOX}/1PWR Asset Management/10_AM spreadsheets"
    items = []
    items += _extract_material_sheet(f"{base}/RET Material Items Database and Requests Log.xlsx",
                                     "RET items database and tracker", 1, 2, 3, 4, 5, "RET")
    items += _extract_material_sheet(f"{base}/FAC Material Items Database and Requests Log.xlsx",
                                     "FAC Material Database and Track", 2, 3, 4, 5, 6, "FAC")
    items += _extract_material_sheet(f"{base}/O&M Material Database.xlsx",
                                     "Sheet1", 3, 4, 5, 6, 7, "O&M")
    items += _extract_material_sheet(f"{base}/General Materials and Items Database.xlsx",
                                     "Sheet1", 2, 3, 4, 5, 6, "General")
    return items


# ── Source 4: Engineering Tools ──────────────────────────────────────

def extract_tools():
    path = f"{DROPBOX}/1PWR Asset Management/10_AM spreadsheets/Engineering Tool List.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["tool_list"]
    items = []
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name:
            continue
        tag = clean(ws.cell(r, 3).value)
        loc = clean(ws.cell(r, 6).value)
        country = CONCESSION_COUNTRY.get(loc, "LSO")
        items.append({
            "name": name,
            "description": clean(ws.cell(r, 4).value),
            "item_class": "FixedAsset",
            "category_name": "Engineering Tools",
            "department_scope": "O&M",
            "quantity": clean_int(ws.cell(r, 5).value) or 1,
            "unit_of_measure": "EA",
            "condition_status": "Good",
            "status": "Available",
            "location_name": loc,
            "country_code": country,
            "legacy_tag": tag,
            "source": "EngineeringToolList",
        })
    return items


# ── Source 5: Meters & Ready Boards → Inventory ─────────────────────

def extract_meters_readyboards():
    path = f"{DROPBOX}/1PWR Asset Management/10_AM spreadsheets/Meters_Meter Enclosures_Ready Boards Database.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)

    items = []

    # Ready boards — individually tracked
    ws = wb["Ready boards stock"]
    for r in range(2, ws.max_row + 1):
        rb_id = clean(ws.cell(r, 1).value)
        if not rb_id:
            continue
        avail = clean(ws.cell(r, 3).value).lower()
        conc = clean(ws.cell(r, 4).value)
        items.append({
            "name": "Ready Board",
            "description": rb_id,
            "item_class": "Inventory",
            "category_name": "Ready Boards",
            "department_scope": "General",
            "serial_number": rb_id,
            "quantity": 1,
            "unit_of_measure": "EA",
            "condition_status": "Good",
            "status": "CheckedOut" if "checked out" in avail else "Available",
            "location_name": conc or "HQ",
            "country_code": CONCESSION_COUNTRY.get(conc, "LSO"),
            "source": "MetersReadyBoards",
        })

    # Meters/enclosures
    ws = wb["Meters_Meter enclosure stock"]
    for r in range(2, ws.max_row + 1):
        enc_id = clean(ws.cell(r, 1).value)
        if not enc_id:
            continue
        enc_type = clean(ws.cell(r, 2).value)
        meter_no = clean(ws.cell(r, 3).value)
        avail = clean(ws.cell(r, 4).value).lower()
        items.append({
            "name": f"Meter Enclosure ({enc_type})" if enc_type else "Meter Enclosure",
            "description": f"Meter: {meter_no}" if meter_no else "",
            "item_class": "Inventory",
            "category_name": "Meters & Enclosures",
            "department_scope": "General",
            "serial_number": enc_id,
            "quantity": 1,
            "unit_of_measure": "EA",
            "condition_status": "Good",
            "status": "CheckedOut" if "checked out" in avail else "Available",
            "country_code": "LSO",
            "source": "MetersReadyBoards",
        })

    return items


# ── Source 6: Miscellaneous Items (2025) ─────────────────────────────

MISC_CATEGORY_CLASS = {
    "Measuring tool components":         "FixedAsset",
    "Welding & Cutting Consumables":     "Consumable",
    "Hand Tools":                        "FixedAsset",
    "Cleaning Products":                 "Consumable",
    "Personal Protective Equipment":     "Consumable",
    "Lubricants & Fluids":               "Consumable",
    "Fasteners & Hardware":              "Material",
    "Safety Equipment":                  "Consumable",
    "Electrical Consumables":            "Consumable",
    "Power Tools & Accessories":         "FixedAsset",
    "Plumbing Fittings":                 "Material",
    "Adhesives & Sealants":              "Consumable",
    "Stationery":                        "Consumable",
    "Abrasives":                         "Consumable",
    "Signage & Labels":                  "Consumable",
    "Painting Supplies":                 "Consumable",
}

def extract_misc_items():
    path = f"{DROPBOX}/1PWR Asset Management/1_Assets Management/Assets Database/Miscellaneous Items From 15-07-2025.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Other Materials"]
    items = []
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 1).value)
        if not name:
            continue
        cat = clean(ws.cell(r, 5).value)
        item_class = MISC_CATEGORY_CLASS.get(cat, "Consumable")
        loc = clean(ws.cell(r, 7).value)
        items.append({
            "name": name,
            "description": clean(ws.cell(r, 2).value),
            "item_class": item_class,
            "category_name": cat or "General",
            "department_scope": "General",
            "purchase_date": clean_date(ws.cell(r, 3).value),
            "quantity": clean_int(ws.cell(r, 8).value) or 1,
            "unit_of_measure": "EA",
            "condition_status": "Good",
            "status": "Available",
            "location_name": loc,
            "assigned_to": clean(ws.cell(r, 6).value),
            "country_code": "LSO",
            "source": "MiscItems2025",
        })
    return items


# ── Source 7: PUECO Stock ────────────────────────────────────────────

def extract_pueco():
    path = f"{DROPBOX}/1PWR Asset Management/1_Assets Management/Assets Database/Pueco Stock.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    items = []
    for sheet_name in ["Fridges", "Air conditioners"]:
        ws = wb[sheet_name]
        header_row = 1
        for r in range(1, 5):
            if clean(ws.cell(r, 2).value) == "Item":
                header_row = r
                break
        for r in range(header_row + 1, ws.max_row + 1):
            name = clean(ws.cell(r, 2).value)
            if not name:
                continue
            serial = clean(ws.cell(r, 4).value)
            cond = clean(ws.cell(r, 5).value)
            loc = clean(ws.cell(r, 7).value)
            avail = clean(ws.cell(r, 9).value).lower()
            country = CONCESSION_COUNTRY.get(loc, "LSO")
            items.append({
                "name": name,
                "description": clean(ws.cell(r, 3).value),
                "item_class": "Inventory",
                "category_name": "PUECO Products",
                "department_scope": "General",
                "serial_number": serial,
                "quantity": 1,
                "unit_of_measure": "EA",
                "condition_status": cond if cond in ("New", "Good", "Fair", "Poor", "Damaged") else "Good",
                "status": "CheckedOut" if "checked out" in avail else "Available",
                "location_name": loc,
                "country_code": country,
                "source": "PUECOStock",
            })
    return items


# ── Source 8: Request logs ───────────────────────────────────────────

def extract_requests():
    base = f"{DROPBOX}/1PWR Asset Management/10_AM spreadsheets"
    requests = []

    sources = [
        (f"{base}/RET Material Items Database and Requests Log.xlsx", "RET items request log", "Material", "RET"),
        (f"{base}/FAC Material Items Database and Requests Log.xlsx", "Form Responses 1", "Material", "FAC"),
        (f"{base}/Meters_Meter Enclosures_Ready Boards Database.xlsx", "ready boards request log", "Inventory", "General"),
    ]

    seq = 1
    for path, sheet, item_class, dept in sources:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[sheet]
        except Exception:
            continue
        for r in range(2, ws.max_row + 1):
            ts = clean(ws.cell(r, 1).value)
            if not ts:
                continue
            requester = clean(ws.cell(r, 3).value)
            description = clean(ws.cell(r, 4).value)
            site = clean(ws.cell(r, 5).value)
            requests.append({
                "request_number": f"REQ-HIST-{seq:04d}",
                "item_class": item_class,
                "department_scope": dept,
                "requested_by_name": requester,
                "requested_by": "",
                "requested_for_country": CONCESSION_COUNTRY.get(site, "LSO"),
                "requested_for_location": site,
                "priority": "Normal",
                "status": "Fulfilled",
                "description": description[:500] if description else f"{item_class} request",
                "requested_date": clean_date(ts),
                "required_date": clean_date(ws.cell(r, 6).value),
                "notes": clean(ws.cell(r, 9).value),
                "source": f"GoogleForm_{dept}",
            })
            seq += 1
    return requests


# ── Categories seed ──────────────────────────────────────────────────

SEED_CATEGORIES = [
    ("FA-VEH", "Vehicles",                 "FixedAsset",  "General", 4,  "DecliningBalance", 0),
    ("FA-MCH", "Machinery & Equipment",    "FixedAsset",  "O&M",    10, "StraightLine", 0),
    ("FA-BLD", "Buildings & Structures",   "FixedAsset",  "General", 10, "StraightLine", 0),
    ("FA-FUR", "Furniture & Fixtures",     "FixedAsset",  "General", 5,  "StraightLine", 0),
    ("FA-ELE", "Electronics & Tools",      "FixedAsset",  "General", 5,  "StraightLine", 0),
    ("FA-IT",  "IT Equipment",             "FixedAsset",  "General", 5,  "StraightLine", 0),
    ("FA-COM", "Communications Equipment", "FixedAsset",  "General", 5,  "StraightLine", 0),
    ("FA-TST", "Test Equipment",           "FixedAsset",  "O&M",    5,  "StraightLine", 0),
    ("FA-PLT", "Plant & Vehicles",         "FixedAsset",  "General", 4,  "DecliningBalance", 0),
    ("FA-ENG", "Engineering Tools",        "FixedAsset",  "O&M",    5,  "StraightLine", 0),
    ("MAT-RET","RET Construction Material","Material",    "RET",    0,  "None", 1),
    ("MAT-FAC","FAC Construction Material","Material",    "FAC",    0,  "None", 1),
    ("MAT-OM", "O&M Spare Parts",          "Material",    "O&M",    0,  "None", 1),
    ("MAT-GEN","General Material",         "Material",    "General", 0,  "None", 1),
    ("MAT-HW", "Fasteners & Hardware",     "Material",    "General", 0,  "None", 1),
    ("CON-PPE","Personal Protective Equip","Consumable",  "O&M",    0,  "None", 1),
    ("CON-OFF","Office Supplies",          "Consumable",  "General", 0,  "None", 1),
    ("CON-CLN","Cleaning Products",        "Consumable",  "General", 0,  "None", 1),
    ("CON-WLD","Welding & Cutting",        "Consumable",  "RET",    0,  "None", 1),
    ("CON-ELC","Electrical Consumables",   "Consumable",  "O&M",    0,  "None", 1),
    ("INV-MTR","Meters & Enclosures",      "Inventory",   "General", 0,  "None", 1),
    ("INV-RDB","Ready Boards",             "Inventory",   "General", 0,  "None", 1),
    ("INV-PUC","PUECO Products",           "Inventory",   "General", 0,  "None", 1),
    ("INV-SPR","Spare Parts & Kits",       "Inventory",   "O&M",    0,  "None", 1),
]

def build_categories():
    return [
        {
            "category_code": code,
            "category_name": name,
            "item_class": ic,
            "department_scope": dept,
            "useful_life_years": ul,
            "depreciation_method": dm,
            "reorder_enabled": re_,
            "description": "",
            "active": 1,
        }
        for code, name, ic, dept, ul, dm, re_ in SEED_CATEGORIES
    ]


# ── Locations seed ───────────────────────────────────────────────────

SEED_LOCATIONS = [
    ("LSO-HQ",  "Headquarters (Maseru)",   "Site",     "LSO", ""),
    ("LSO-MAT", "Matsieng Concession",     "Site",     "LSO", ""),
    ("LSO-TLH", "Thaba-Liholo Concession", "Site",     "LSO", ""),
    ("LSO-MAS", "Mashai Concession",       "Site",     "LSO", ""),
    ("LSO-SHG", "Sehong Concession",       "Site",     "LSO", ""),
    ("LSO-LEB", "Lebakeng Concession",     "Site",     "LSO", ""),
    ("LSO-RIB", "Ribaneng Concession",     "Site",     "LSO", ""),
    ("LSO-KET", "Ketane Concession",       "Site",     "LSO", ""),
    ("LSO-SEH", "Sehonghong Concession",   "Site",     "LSO", ""),
    ("LSO-LSB", "Letšeng-la-Baroa",        "Site",     "LSO", ""),
    ("LSO-MAK", "Makhoakhoeng",            "Site",     "LSO", ""),
    ("LSO-TOS", "Tosing",                  "Site",     "LSO", ""),
    ("LSO-SEB", "Sebelekoane",             "Site",     "LSO", ""),
    ("ZMB-HQ",  "Zambia HQ",              "Site",     "ZMB", ""),
    ("BEN-HQ",  "Benin HQ",               "Site",     "BEN", ""),
]

def build_locations():
    return [
        {
            "location_code": code,
            "location_name": name,
            "location_type": lt,
            "country_code": cc,
            "parent_location_code": parent,
            "active": 1,
        }
        for code, name, lt, cc, parent in SEED_LOCATIONS
    ]


# ── Countries seed ───────────────────────────────────────────────────

SEED_COUNTRIES = [
    ("1", "LSO", "Lesotho",  1),
    ("2", "ZMB", "Zambia",   1),
    ("3", "BEN", "Benin",    1),
]

def build_countries():
    return [
        {"country_id": cid, "country_code": cc, "country_name": cn, "active": a}
        for cid, cc, cn, a in SEED_COUNTRIES
    ]


# ── Normalise & merge ────────────────────────────────────────────────

def category_code_for(item):
    ic = item.get("item_class", "")
    cat = item.get("category_name", "")
    if item.get("category_code"):
        return item["category_code"]
    for code, name, cls, *_ in SEED_CATEGORIES:
        if cls == ic and (cat.lower() in name.lower() or name.lower() in cat.lower()):
            return code
    defaults = {"FixedAsset": "FA-MCH", "Material": "MAT-GEN", "Consumable": "CON-OFF", "Inventory": "INV-SPR"}
    return defaults.get(ic, "MAT-GEN")

def location_code_for(item):
    loc = (item.get("location_name") or "").upper().strip()
    for code, name, *_ in SEED_LOCATIONS:
        short = code.split("-")[-1]
        if short == loc or loc in name.upper():
            return code
    return "LSO-HQ" if item.get("country_code") == "LSO" else f"{item.get('country_code', 'LSO')}-HQ"

def normalise_item(item, seq_counters):
    ic = item["item_class"]
    cc = item.get("country_code", "LSO")
    cat_code = category_code_for(item)
    loc_code = location_code_for(item)

    key = f"{ic}:{cc}"
    seq_counters[key] = seq_counters.get(key, 0) + 1
    seq = seq_counters[key]

    prefix_map = {"FixedAsset": "FA", "Material": "MAT", "Consumable": "CON", "Inventory": "INV"}
    prefix = prefix_map.get(ic, "ITM")
    asset_tag = f"1PWR-{prefix}-{cc}-{seq:06d}"

    record = {
        "name": item["name"],
        "description": item.get("description", ""),
        "item_class": ic,
        "category_code": cat_code,
        "country_code": cc,
        "location_code": loc_code,
        "asset_tag": asset_tag,
        "serial_number": item.get("serial_number", ""),
        "manufacturer": item.get("manufacturer", ""),
        "model": item.get("model", ""),
        "purchase_date": item.get("purchase_date", ""),
        "purchase_price": item.get("purchase_price", 0),
        "current_value": item.get("current_value", 0),
        "salvage_value": item.get("salvage_value", 0),
        "condition_status": item.get("condition_status", "Good"),
        "status": item.get("status", "Available"),
        "quantity": item.get("quantity", 1),
        "unit_of_measure": item.get("unit_of_measure", "EA"),
        "notes": item.get("notes", ""),
        "department_scope": item.get("department_scope", "General"),
        "legacy_tag": item.get("legacy_tag", ""),
        "source": item.get("source", ""),
        "created_at": NOW,
        "updated_at": NOW,
    }
    return record


# ── Deduplication ────────────────────────────────────────────────────

def deduplicate(items):
    seen = {}
    out = []
    for item in items:
        name = item.get("name", "").lower().strip()
        serial = item.get("serial_number", "").lower().strip()
        tag = item.get("legacy_tag", "").lower().strip()
        source = item.get("source", "")

        key = dedup_key(name, serial, tag) if (serial or tag) else None

        if key and key in seen:
            existing = seen[key]
            priority = ["FY2025_PPE_Register", "AssetSpreadsheetDB"]
            if source in priority and existing["source"] not in priority:
                seen[key] = item
                for i, o in enumerate(out):
                    if dedup_key(o["name"].lower(), o.get("serial_number", "").lower(), o.get("legacy_tag", "").lower()) == key:
                        out[i] = item
                        break
            continue

        if key:
            seen[key] = item
        out.append(item)
    return out


# ── PPE enrichment ───────────────────────────────────────────────────

def enrich_with_ppe(assets, ppe_items):
    ppe_by_tag = {}
    for p in ppe_items:
        for t in p.get("legacy_tags", []):
            ppe_by_tag[t.lower()] = p

    enriched = 0
    for a in assets:
        tag = (a.get("legacy_tag") or "").lower()
        if tag and tag in ppe_by_tag:
            ppe = ppe_by_tag[tag]
            if ppe["purchase_price"] > 0 and a.get("purchase_price", 0) == 0:
                a["purchase_price"] = ppe["purchase_price"]
            if ppe.get("current_value", 0) > 0:
                a["current_value"] = ppe["current_value"]
            if ppe.get("salvage_value", 0) > 0:
                a["salvage_value"] = ppe["salvage_value"]
            enriched += 1
    return enriched


# ── Main ─────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  1PWR Asset Management -- Data Migration ETL")
    print("=" * 60)

    # Extract
    print("\n[1/8] Extracting Asset Spreadsheet Database...")
    asset_db = extract_asset_db()
    print(f"       → {len(asset_db)} items")

    print("[2/8] Extracting FY2025 PPE Register...")
    ppe_items = extract_ppe_register()
    print(f"       → {len(ppe_items)} financial records")

    print("[3/8] Extracting Material databases (RET, FAC, O&M, General)...")
    materials = extract_materials()
    print(f"       → {len(materials)} material items")

    print("[4/8] Extracting Engineering Tools...")
    tools = extract_tools()
    print(f"       → {len(tools)} tools")

    print("[5/8] Extracting Meters & Ready Boards...")
    meters_rb = extract_meters_readyboards()
    print(f"       → {len(meters_rb)} units")

    print("[6/8] Extracting Misc Items (2025)...")
    misc = extract_misc_items()
    print(f"       → {len(misc)} items")

    print("[7/8] Extracting PUECO Stock...")
    pueco = extract_pueco()
    print(f"       → {len(pueco)} items")

    print("[8/8] Extracting historical requests...")
    requests = extract_requests()
    print(f"       → {len(requests)} requests")

    # Enrich asset_db with PPE financial data
    print("\n[Enrich] Cross-referencing PPE register with asset database...")
    enriched = enrich_with_ppe(asset_db, ppe_items)
    print(f"         → {enriched} assets enriched with financial data")

    # Merge all items
    all_items = asset_db + materials + tools + meters_rb + misc + pueco
    print(f"\n[Merge]  Total raw items: {len(all_items)}")

    # Deduplicate
    deduped = deduplicate(all_items)
    print(f"[Dedup]  After deduplication: {len(deduped)}")

    # Normalise
    print("[Norm]   Normalising and assigning asset tags...")
    seq_counters = {}
    normalised = [normalise_item(item, seq_counters) for item in deduped]

    # Stats
    by_class = {}
    by_source = {}
    by_country = {}
    for item in normalised:
        ic = item["item_class"]
        by_class[ic] = by_class.get(ic, 0) + 1
        src = item["source"]
        by_source[src] = by_source.get(src, 0) + 1
        cc = item["country_code"]
        by_country[cc] = by_country.get(cc, 0) + 1

    print("\n" + "─" * 40)
    print("  MIGRATION SUMMARY")
    print("─" * 40)
    print(f"\n  Total items to migrate: {len(normalised)}")
    print(f"\n  By item class:")
    for k in sorted(by_class):
        print(f"    {k:15s} {by_class[k]:>6d}")
    print(f"\n  By source:")
    for k in sorted(by_source):
        print(f"    {k:25s} {by_source[k]:>6d}")
    print(f"\n  By country:")
    for k in sorted(by_country):
        print(f"    {k:5s} {by_country[k]:>6d}")
    print(f"\n  Historical requests: {len(requests)}")

    # Write output
    categories = build_categories()
    locations = build_locations()
    countries = build_countries()

    outputs = {
        "am_core_assets.json": normalised,
        "pr_master_categories.json": categories,
        "pr_master_locations.json": locations,
        "pr_master_countries.json": countries,
        "pr_master_requests.json": requests,
        "ppe_financial_data.json": ppe_items,
    }

    print(f"\n[Write]  Writing output files to {OUT_DIR}/")
    for filename, data in outputs.items():
        out_path = OUT_DIR / filename
        with open(out_path, "w") as f:
            json.dump(data, f, indent=2, default=str)
        print(f"         {filename:35s} ({len(data):>5d} records, {os.path.getsize(out_path):>10,d} bytes)")

    print("\n✓ ETL complete. Review JSON files before importing to Firestore.")
    print(f"  Output directory: {OUT_DIR}")

if __name__ == "__main__":
    main()
